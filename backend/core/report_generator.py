"""
Advanced Excel Report Generator for Smart Competitor Finder
Generates comprehensive Excel reports with keyword, semantic, and sector analysis data
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
from typing import List, Dict, Any
import os


class ReportGenerator:
    """Advanced Excel report generator with professional formatting and multiple sheets"""
    
    def __init__(self):
        self.workbook = None
        self.styles = self._init_styles()
    
    def _init_styles(self) -> Dict[str, Any]:
        """Initialize Excel styling configurations"""
        return {
            'header': {
                'font': Font(bold=True, color='FFFFFF', size=12),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'subheader': {
                'font': Font(bold=True, color='366092', size=11),
                'fill': PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'data': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
            },
            'number': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'number_format': '0.0%'
            },
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def generate_comprehensive_report(
        self,
        client_url: str,
        client_keywords: List[str],
        analysis_results: List[Dict],
        output_path: str = None,
        failed_sites: List[Dict] = None  # 🆕 FASE 1: Failed sites for dedicated sheet
    ) -> str:
        """
        Generate comprehensive Excel report with multiple sheets
        
        Args:
            client_url: The analyzed client website URL
            client_keywords: Selected keywords for analysis
            analysis_results: List of competitor analysis results
            output_path: Custom output file path
            failed_sites: List of failed sites with error details (FASE 1)
            
        Returns:
            str: Path to generated Excel file
        """
        # Create workbook
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in [ws.title for ws in self.workbook.worksheets]:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Generate different sheets
        self._create_summary_sheet(client_url, client_keywords, analysis_results)
        self._create_detailed_results_sheet(analysis_results)
        self._create_sector_analysis_sheet(analysis_results)
        self._create_keyword_analysis_sheet(analysis_results, client_keywords)
        self._create_semantic_analysis_sheet(analysis_results)
        
        # 🆕 FASE 1: Create failed sites sheet if there are failures
        if failed_sites and len(failed_sites) > 0:
            self._create_failed_sites_sheet(failed_sites)
        
        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"reports/competitor_analysis_{timestamp}.xlsx"
        
        # Create reports directory if it doesn't exist
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save workbook
        self.workbook.save(output_path)
        
        return output_path
    
    def _create_summary_sheet(
        self,
        client_url: str,
        client_keywords: List[str],
        analysis_results: List[Dict]
    ):
        """Create executive summary sheet"""
        ws = self.workbook.create_sheet("Executive Summary", 0)
        
        # Title and metadata
        ws['A1'] = "Smart Competitor Finder - Analysis Report"
        ws['A1'].font = Font(bold=True, size=16, color='366092')
        ws.merge_cells('A1:F1')
        
        # 🆕 Data + Ora nel report
        current_datetime = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        
        ws['A3'] = f"Client Website: {client_url}"
        ws['A4'] = f"Report generato il: {current_datetime}"
        ws['A5'] = f"Keywords Analyzed: {', '.join(client_keywords)}"
        ws['A6'] = f"Total Competitors Analyzed: {len(analysis_results)}"
        
        # Summary statistics basate su KPI classification
        direct_competitors = [r for r in analysis_results if r.get('competitor_status', {}).get('category') == 'DIRECT']
        potential_competitors = [r for r in analysis_results if r.get('competitor_status', {}).get('category') == 'POTENTIAL']
        non_competitors = [r for r in analysis_results if r.get('competitor_status', {}).get('category') == 'NON_COMPETITOR']
        
        ws['A8'] = "ANALISI PER CATEGORIA KPI"
        ws['A8'].font = Font(bold=True, size=14, color='366092')
        
        summary_data = [
            ["Categoria", "Numero", "Percentuale"],
            ["🟢 Competitor Diretti (≥60%)", len(direct_competitors), f"{len(direct_competitors) / len(analysis_results) * 100:.1f}%" if analysis_results else "0%"],
            ["🟡 Da Valutare (40-59%)", len(potential_competitors), f"{len(potential_competitors) / len(analysis_results) * 100:.1f}%" if analysis_results else "0%"],
            ["🔴 Non Competitor (<40%)", len(non_competitors), f"{len(non_competitors) / len(analysis_results) * 100:.1f}%" if analysis_results else "0%"],
            ["", "", ""],
            ["Score Medio", f"{sum(r.get('score', 0) for r in analysis_results) / len(analysis_results):.1f}%" if analysis_results else "0%", ""],
        ]
        
        for row_idx, row_data in enumerate(summary_data, 10):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 10:  # Header row
                    cell.font = self.styles['header']['font']
                    cell.fill = self.styles['header']['fill']
                    cell.alignment = self.styles['header']['alignment']
                cell.border = self.styles['border']
        
        # Top 50 competitors ordinati per score con classe KPI
        ws['A16'] = "TOP 50 COMPETITOR PER CATEGORIA"
        ws['A16'].font = Font(bold=True, size=14, color='366092')
        
        # Sort by score and take top 50
        top_competitors = sorted(
            analysis_results,
            key=lambda x: x.get('score', 0),
            reverse=True
        )[:50]
        
        headers = ["Rank", "Website", "Score", "Categoria KPI", "Azione Consigliata"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=18, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        for row_idx, competitor in enumerate(top_competitors, 19):
            status = competitor.get('competitor_status', {})
            row_data = [
                row_idx - 18,
                competitor.get('url', 'N/A'),
                f"{competitor.get('score', 0):.1f}%",
                f"{status.get('emoji', '⚪')} {status.get('label', 'Non classificato')}",
                status.get('action', 'N/A')
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles['data']['font']
                cell.alignment = self.styles['data']['alignment']
                cell.border = self.styles['border']
                
                # Colora riga in base a categoria
                if status.get('color') == 'green':
                    cell.fill = PatternFill(start_color='E6F9E6', end_color='E6F9E6', fill_type='solid')
                elif status.get('color') == 'yellow':
                    cell.fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
                elif status.get('color') == 'red':
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_detailed_results_sheet(self, analysis_results: List[Dict]):
        """Create detailed results sheet with all competitor data and KPI classification"""
        ws = self.workbook.create_sheet("Detailed Results")
        
        # Headers con classificazione KPI (senza Score %)
        headers = [
            "URL", "Categoria KPI", "Azione", 
            "Keywords Found", "Keyword Count", "Title"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        # Data rows con classificazione KPI (senza Score %)
        for row_idx, result in enumerate(analysis_results, 2):
            status = result.get('competitor_status', {})
            row_data = [
                result.get('url', 'N/A'),
                f"{status.get('emoji', '⚪')} {status.get('label', 'Non classificato')}",
                status.get('action', 'N/A'),
                ', '.join(result.get('keywords_found', [])),
                len(result.get('keywords_found', [])),
                result.get('title', 'N/A')
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.styles['border']
                
                # Applica colore di sfondo in base a categoria KPI
                if status.get('color') == 'green':
                    cell.fill = PatternFill(start_color='E6F9E6', end_color='E6F9E6', fill_type='solid')
                elif status.get('color') == 'yellow':
                    cell.fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
                elif status.get('color') == 'red':
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                
                # Formatting specifico per colonne (senza Score %)
                if col_idx == 5:  # Count column (ora colonna 5 invece di 6)
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.font = self.styles['data']['font']
                    cell.alignment = self.styles['data']['alignment']
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_sector_analysis_sheet(self, analysis_results: List[Dict]):
        """Create KPI category distribution analysis sheet"""
        ws = self.workbook.create_sheet("Analisi KPI")
        
        # Raggruppa per categoria KPI
        direct = [r for r in analysis_results if r.get('competitor_status', {}).get('category') == 'DIRECT']
        potential = [r for r in analysis_results if r.get('competitor_status', {}).get('category') == 'POTENTIAL']
        non_comp = [r for r in analysis_results if r.get('competitor_status', {}).get('category') == 'NON_COMPETITOR']
        
        # Create KPI summary table
        ws['A1'] = "ANALISI DISTRIBUZIONE PER CATEGORIA KPI"
        ws['A1'].font = Font(bold=True, size=14, color='366092')
        ws.merge_cells('A1:E1')
        
        headers = ["Categoria", "Emoji", "Numero", "Percentuale", "Score Medio"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        total = len(analysis_results) if analysis_results else 1
        
        categories_data = [
            ("Competitor Diretti (≥60%)", "🟢", direct, 'E6F9E6'),
            ("Da Valutare (40-59%)", "🟡", potential, 'FFF9E6'),
            ("Non Competitor (<40%)", "🔴", non_comp, 'FFE6E6')
        ]
        
        row_idx = 4
        for category_name, emoji, competitors, bg_color in categories_data:
            count = len(competitors)
            percentage = (count / total) * 100
            avg_score = sum(c.get('score', 0) for c in competitors) / count if count > 0 else 0
            
            row_data = [
                category_name,
                emoji,
                count,
                f"{percentage:.1f}%",
                f"{avg_score:.1f}%"
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.styles['border']
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.font = self.styles['data']['font']
                cell.alignment = self.styles['data']['alignment']
            
            row_idx += 1
        
        # Aggiungi lista dettagliata per ogni categoria
        row_idx += 2
        
        for category_name, emoji, competitors, bg_color in categories_data:
            if not competitors:
                continue
                
            ws.cell(row=row_idx, column=1, value=f"{emoji} {category_name}").font = Font(bold=True, size=12)
            row_idx += 1
            
            # Headers per lista
            sub_headers = ["URL", "Score", "Keywords"]
            for col_idx, header in enumerate(sub_headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=header)
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.border = self.styles['border']
            
            row_idx += 1
            
            # Competitor list - top 30 per categoria
            for comp in sorted(competitors, key=lambda x: x.get('score', 0), reverse=True)[:30]:
                cell_data = [
                    comp.get('url', 'N/A'),
                    f"{comp.get('score', 0):.1f}%",
                    ', '.join(comp.get('keywords_found', []))[:50] + "..."
                ]
                
                for col_idx, value in enumerate(cell_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.styles['border']
                    cell.font = Font(size=9)
                
                row_idx += 1
            
            row_idx += 1  # Spazio tra categorie
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_keyword_analysis_sheet(self, analysis_results: List[Dict], client_keywords: List[str]):
        """Create keyword analysis breakdown sheet"""
        ws = self.workbook.create_sheet("Keyword Analysis")
        
        # Keyword frequency analysis
        keyword_frequency = {}
        for result in analysis_results:
            for keyword in result.get('keywords_found', []):
                if keyword.lower() in [k.lower() for k in client_keywords]:
                    if keyword not in keyword_frequency:
                        keyword_frequency[keyword] = 0
                    keyword_frequency[keyword] += 1
        
        ws['A1'] = "KEYWORD FREQUENCY ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14, color='366092')
        
        headers = ["Keyword", "Frequency", "Percentage"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        total_competitors = len(analysis_results)
        row_idx = 4
        
        for keyword in sorted(keyword_frequency.keys(), key=keyword_frequency.get, reverse=True):
            frequency = keyword_frequency[keyword]
            percentage = frequency / total_competitors if total_competitors > 0 else 0
            
            row_data = [keyword, frequency, percentage]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.styles['border']
                
                if col_idx == 3:  # Percentage
                    cell.number_format = '0.0%'
                    cell.alignment = self.styles['number']['alignment']
                else:
                    cell.font = self.styles['data']['font']
                    cell.alignment = self.styles['data']['alignment']
            
            row_idx += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_semantic_analysis_sheet(self, analysis_results: List[Dict]):
        """Create semantic analysis details sheet"""
        ws = self.workbook.create_sheet("Semantic Analysis")
        
        ws['A1'] = "SEMANTIC SIMILARITY ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14, color='366092')
        
        # Filter results that have semantic analysis data
        semantic_results = [r for r in analysis_results if r.get('semantic_similarity') is not None]
        
        if not semantic_results:
            ws['A3'] = "No semantic analysis data available"
            return
        
        headers = ["URL", "Semantic Score", "Sector Match", "Content Summary"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        # Sort by semantic score
        semantic_results.sort(key=lambda x: x.get('semantic_similarity', 0), reverse=True)
        
        for row_idx, result in enumerate(semantic_results, 4):
            row_data = [
                result.get('url', 'N/A'),
                result.get('semantic_similarity', 0),
                "Yes" if result.get('is_relevant', True) else "No",
                result.get('content_summary', 'N/A')[:100] + "..." if len(result.get('content_summary', '')) > 100 else result.get('content_summary', 'N/A')
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.styles['border']
                
                if col_idx == 2:  # Semantic score
                    cell.number_format = '0.0%'
                    cell.alignment = self.styles['number']['alignment']
                else:
                    cell.font = self.styles['data']['font']
                    cell.alignment = self.styles['data']['alignment']
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_failed_sites_sheet(self, failed_sites: List[Dict]):
        """
        🚨 FASE 1: Create sheet for failed sites with error details and suggestions
        
        Args:
            failed_sites: List of dicts with keys: url, error, suggestion, timestamp
        """
        ws = self.workbook.create_sheet("Siti Non Analizzati")
        
        # Headers with red styling (alert color)
        headers = ['URL', 'Motivo Errore', 'Suggerimento', 'Timestamp']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF', size=12)
            cell.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')  # Bootstrap danger red
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.styles['border']
        
        # Data rows
        for row_idx, site in enumerate(failed_sites, start=2):
            # URL column
            url_cell = ws.cell(row=row_idx, column=1, value=site.get('url', 'N/A'))
            url_cell.font = Font(size=10, underline='single', color='0563C1')  # Link style
            url_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            url_cell.border = self.styles['border']
            
            # Error column
            error_cell = ws.cell(row=row_idx, column=2, value=site.get('error', 'Errore sconosciuto'))
            error_cell.font = Font(size=10)
            error_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            error_cell.border = self.styles['border']
            
            # Suggestion column (bold for emphasis)
            suggestion_cell = ws.cell(row=row_idx, column=3, value=site.get('suggestion', 'Verifica manualmente'))
            suggestion_cell.font = Font(size=10, bold=True, color='856404')  # Warning brown
            suggestion_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            suggestion_cell.border = self.styles['border']
            
            # Timestamp column
            timestamp_cell = ws.cell(row=row_idx, column=4, value=site.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            timestamp_cell.font = Font(size=9, color='6C757D')  # Muted gray
            timestamp_cell.alignment = Alignment(horizontal='center', vertical='center')
            timestamp_cell.border = self.styles['border']
        
        # Auto-size columns with max width constraint
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            for cell in col:
                try:
                    if cell.value:
                        # Handle wrapped text by checking actual content length
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width with reasonable constraints
            adjusted_width = min(max(max_length + 2, 15), 60)  # Min 15, Max 60
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add note at the bottom
        note_row = len(failed_sites) + 3
        ws.merge_cells(f'A{note_row}:D{note_row}')
        note_cell = ws[f'A{note_row}']
        note_cell.value = "ℹ️ Nota: Questi siti non sono stati analizzati a causa degli errori elencati. Si consiglia di verificarli manualmente o contattare il supporto tecnico."
        note_cell.font = Font(size=9, italic=True, color='6C757D')
        note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Summary stats
        summary_row = note_row + 2
        ws[f'A{summary_row}'] = f"Totale siti falliti: {len(failed_sites)}"
        ws[f'A{summary_row}'].font = Font(bold=True, size=11, color='DC3545')


def create_sample_report():
    """Create a sample report for testing"""
    generator = ReportGenerator()
    
    # Sample data
    client_url = "https://example-agency.com"
    client_keywords = ["digital marketing", "web design", "SEO", "advertising"]
    
    sample_results = [
        {
            "url": "https://competitor1.com",
            "total_score": 0.85,
            "keyword_score": 0.75,
            "semantic_score": 0.90,
            "sector": "Digital Marketing",
            "is_relevant": True,
            "keywords_found": ["digital marketing", "SEO"],
            "semantic_similarity": 0.90,
            "relevance_label": "Highly Relevant",
            "analysis_notes": "Strong competitor in digital marketing space"
        },
        {
            "url": "https://competitor2.com",
            "total_score": 0.45,
            "keyword_score": 0.30,
            "semantic_score": 0.55,
            "sector": "Web Development",
            "is_relevant": True,
            "keywords_found": ["web design"],
            "semantic_similarity": 0.55,
            "relevance_label": "Moderately Relevant",
            "analysis_notes": "Focuses more on development than marketing"
        }
    ]
    
    return generator.generate_comprehensive_report(
        client_url, client_keywords, sample_results, "sample_report.xlsx"
    )


if __name__ == "__main__":
    # Test the report generator
    report_path = create_sample_report()
    print(f"Sample report generated: {report_path}")